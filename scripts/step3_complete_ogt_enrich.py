"""
Step ③ (Complete): OGT substrate enrichment analysis with 6,777 DEGs
"""
import csv, json, os
from pathlib import Path
from collections import Counter
from scipy.stats import fisher_exact
import openpyxl

# ── Load OGT interactors ──────────────────────────────────────────
wb = openpyxl.load_workbook(
    r"D:\Cardiac_Ogt_MultiOmics\validation_data\OGT-PIN_S1.xlsx",
    read_only=True
)
ws = wb.active
hdr = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}

ogt_high = set()
for r in range(2, ws.max_row + 1):
    gene = ws.cell(r, hdr["gene_name_b"]).value
    taxon = ws.cell(r, hdr["ncbi_taxonomy identifier_b"]).value
    hs = ws.cell(r, hdr["High-stringency interactor"]).value
    if gene and taxon and "Homo sapiens" in str(taxon):
        if hs and "yes" in str(hs).lower():
            ogt_high.add(str(gene).strip())

print(f"OGT high-stringency interactors: {len(ogt_high)}")

# ── Load 6,777 DEGs from GSE193997 ────────────────────────────────
deg_path = r"D:\miri_paper\results\tables\GSE193997_DESeq2_Sham_vs_IR6h.csv"
deg_genes = []
deg_data = {}

with open(deg_path) as f:
    reader = csv.DictReader(f)
    for row in reader:
        try:
            padj = float(row["padj"])
            if padj < 0.05:
                gene = row["symbol"].strip()
                deg_genes.append(gene)
                deg_data[gene] = {
                    "log2FC": float(row["log2FoldChange"]),
                    "padj": padj,
                    "baseMean": float(row["baseMean"]),
                }
        except (ValueError, KeyError):
            pass

deg_set = set(deg_genes)
print(f"DEGs (padj < 0.05): {len(deg_set)}")

# ── Intersection ──────────────────────────────────────────────────
overlap = deg_set & ogt_high
print(f"\n{'='*60}")
print(f"OGT-PIN x DEG Overlap: {len(overlap)} genes")
print(f"{'='*60}")

# Fisher exact test
n_deg = len(deg_set)
n_ogt = len(ogt_high)
n_overlap = len(overlap)
n_bg = 20000

a = n_overlap
b = n_deg - a
c = n_ogt - a
d = n_bg - n_deg - c

or_val, p_val = fisher_exact([[a, b], [c, max(d, 1)]], alternative="greater")
print(f"\nFisher exact test (background = {n_bg} protein-coding genes):")
print(f"  DEGs that are OGT interactors: {a}/{n_deg} ({100*a/n_deg:.1f}%)")
print(f"  Background OGT interactors: {n_ogt}/{n_bg} ({100*n_ogt/n_bg:.1f}%)")
print(f"  OR = {or_val:.2f}, P = {p_val:.8f}")
if p_val < 0.05:
    print(f"  >>> SIGNIFICANT enrichment of OGT interactors in DEGs")
else:
    print(f"  >>> NOT significant")

# ── Overlap gene details ─────────────────────────────────────────
print(f"\n--- Overlapping Genes (sorted by log2FC) ---")
print(f"{'Gene':12s} {'log2FC':>8s}  {'padj':>10s}  {'Direction':>10s}")
print("-" * 48)
overlap_sorted = sorted(overlap, key=lambda g: deg_data[g]["log2FC"])
up_count = 0
down_count = 0
for g in overlap_sorted:
    d = deg_data[g]
    direction = "UP" if d["log2FC"] > 0 else "DOWN"
    if d["log2FC"] > 0: up_count += 1
    else: down_count += 1
    print(f"  {g:12s} {d['log2FC']:+8.3f}  {d['padj']:10.2e}  {direction:>10s}")

print(f"\n  UP: {up_count}, DOWN: {down_count}")

# ── WGCNA module mapping ─────────────────────────────────────────
wgcna_path = r"D:\miri_paper\results\tables\WGCNA_module_assignments.csv"
if os.path.exists(wgcna_path):
    print(f"\n--- WGCNA Module Distribution of Overlap Genes ---")
    with open(wgcna_path) as f:
        reader = csv.DictReader(f)
        wgcna = {}
        for row in reader:
            g = row.get("gene", row.get("symbol", "")).strip()
            mod = row.get("module", row.get("Module", "")).strip()
            if g and mod:
                wgcna[g] = mod

    mod_counter = Counter()
    mapped = 0
    for g in overlap_sorted:
        mod = wgcna.get(g, "not_in_WGCNA")
        mod_counter[mod] += 1
        if mod != "not_in_WGCNA":
            mapped += 1
            print(f"  {g:12s}  -> {mod}")

    print(f"\n  Mapped: {mapped}/{len(overlap)}")
    print(f"  Module distribution: {mod_counter.most_common(10)}")

# ── Key pathway genes check ──────────────────────────────────────
print(f"\n--- Key Nrf2/Ferroptosis Genes (are they in OGT-PIN?) ---")
key_genes = [
    ("NFE2L2", "NRF2 master TF"),
    ("KEAP1", "NRF2 inhibitor"),
    ("BACH1", "NRF2 competitor"),
    ("HMOX1", "Ferroptosis / heme degradation"),
    ("SLC7A11", "Ferroptosis / cystine import"),
    ("GPX4", "Ferroptosis / lipid peroxide scavenger"),
    ("FTH1", "Iron storage"),
    ("TFRC", "Iron import"),
    ("NQO1", "NRF2 target / antioxidant"),
    ("GCLC", "Glutathione synthesis"),
    ("RELA", "NF-kB p65"),
    ("NFKB1", "NF-kB subunit"),
    ("STAT3", "STAT3 TF"),
    ("HIF1A", "HIF-1 alpha"),
]
for gene, desc in key_genes:
    in_ogt = "YES" if gene in ogt_high else "-"
    in_deg = f"DEG (log2FC={deg_data[gene]['log2FC']:+.2f})" if gene in deg_data else "-"
    print(f"  {gene:10s}  OGT-PIN: {in_ogt:4s}  {in_deg}  ({desc})")

print(f"\n=== DONE ===")
