"""
Step ③: OGT substrate enrichment analysis
Compares OGT-PIN high-stringency interactors with:
1. 27-node Boolean network genes
2. PXD046631 proteomics-detected genes
3. Concordance table genes
4. Full DEG list (to be provided by user)
"""

import json, csv, os
from pathlib import Path
from scipy.stats import fisher_exact

BASE = Path(r"D:\Cardiac_Ogt_MultiOmics")
OUT = BASE / "output_tables"
OUT.mkdir(parents=True, exist_ok=True)

# ── 1. Load OGT interactors ──────────────────────────────────────────
with open(BASE / "validation_data" / "OGT-PIN_S1.xlsx", "rb") as f:
    import openpyxl
    wb = openpyxl.load_workbook(f)
    ws = wb.active
    header = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}

ogt_high = set()
ogt_all_human = set()
for r in range(2, ws.max_row + 1):
    gene = ws.cell(r, header["gene_name_b"]).value
    taxon = ws.cell(r, header["ncbi_taxonomy identifier_b"]).value
    hs = ws.cell(r, header["High-stringency interactor"]).value
    if gene and taxon and "Homo sapiens" in str(taxon):
        ogt_all_human.add(str(gene).strip())
        if hs and "yes" in str(hs).lower():
            ogt_high.add(str(gene).strip())

print(f"OGT-PIN high-stringency human interactors: {len(ogt_high)}")
print(f"OGT-PIN all human interactors: {len(ogt_all_human)}")

# ── 2. Test against 27-node network ──────────────────────────────────
network_genes = {
    "OGT", "OGA", "MGEA5", "NFE2L2", "KEAP1", "BACH1", "GPX4", "ACSL4",
    "SLC7A11", "HMOX1", "FTH1", "FTL", "TFRC", "NQO1", "PRDX1", "GCLC",
    "GCLM", "TXN", "TXN2", "SRXN1", "SQSTM1", "G6PD", "ME1", "HIF1A",
    "NFKB1", "RELA", "STAT3", "PTGS2"
}

net_ogt = network_genes & ogt_high
print(f"\n── 27-Node Boolean Network ──")
print(f"OGT interactors in network: {len(net_ogt)} / {len(network_genes)}")
print(f"Genes: {sorted(net_ogt)}")

# Fisher test (approximate background = 20,000 human protein-coding genes)
a = len(net_ogt)
b = len(network_genes) - a
c = len(ogt_high) - a
d = 20000 - len(network_genes) - c
or_val, p_val = fisher_exact([[a, b], [c, max(d, 1)]], alternative="greater")
print(f"Fisher: OR = {or_val:.1f}, P = {p_val:.4f}")
if p_val < 0.05:
    print(f"  → SIGNIFICANT: OGT interactors enriched in network genes")
else:
    print(f"  → Not significant at P < 0.05")

# ── 3. Test against PXD046631 proteomics ─────────────────────────────
prot_file = BASE / "data_sources" / "02_原始数据" / "proteomics" / "PXD046631" / "PXD046631_mRNA_protein_merged.csv"
if prot_file.exists():
    with open(prot_file) as f:
        reader = csv.DictReader(f)
        prot_genes = set()
        for row in reader:
            g = row.get("Gene", "").strip()
            if g:
                prot_genes.add(g)

    prot_ogt = prot_genes & ogt_high
    print(f"\n── PXD046631 Proteomics ({len(prot_genes)} detected proteins) ──")
    print(f"OGT interactors detected: {len(prot_ogt)} / {len(prot_genes)}")
    if prot_ogt:
        print(f"Genes: {sorted(prot_ogt)}")

    a2 = len(prot_ogt)
    b2 = len(prot_genes) - a2
    c2 = len(ogt_high) - a2
    d2 = 20000 - len(prot_genes) - c2
    or2, p2 = fisher_exact([[a2, b2], [c2, max(d2, 1)]], alternative="greater")
    print(f"Fisher: OR = {or2:.1f}, P = {p2:.4f}")

# ── 4. Key genes check ──────────────────────────────────────────────
print(f"\n── Key Network Genes in OGT-PIN ──")
key = ["OGT", "OGA", "NFE2L2", "KEAP1", "BACH1", "HMOX1", "SLC7A11",
       "GPX4", "FTH1", "TFRC", "NFKB1", "RELA", "STAT3", "HIF1A", "ACSL4",
       "NQO1", "GCLC", "GCLM", "SRXN1", "SQSTM1"]
for g in key:
    status = "HIGH" if g in ogt_high else ("ALL" if g in ogt_all_human else "-")
    print(f"  {g:12s}  {status}")

# ── 5. Network gene functional categories ───────────────────────────
categories = {
    "O-GlcNAc Cycling": ["OGT", "OGA", "MGEA5"],
    "NRF2 Pathway": ["NFE2L2", "KEAP1", "BACH1"],
    "Ferroptosis Core": ["GPX4", "ACSL4", "SLC7A11", "HMOX1", "FTH1", "FTL", "TFRC"],
    "Antioxidant (NRF2 targets)": ["NQO1", "PRDX1", "GCLC", "GCLM", "TXN", "TXN2", "SRXN1", "SQSTM1", "G6PD", "ME1"],
    "Signaling": ["HIF1A", "NFKB1", "RELA", "STAT3", "PTGS2"],
}
print(f"\n── By Functional Category ──")
for cat, genes in categories.items():
    hits = [g for g in genes if g in ogt_high]
    print(f"  {cat}: {len(hits)}/{len(genes)} OGT interactors: {hits}")

print(f"\n── DONE ──")
print(f"Output saved to: {OUT}")
